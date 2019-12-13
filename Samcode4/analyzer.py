import pickle as p
import openpyxl as op
wb_obj = op.load_workbook("Distances.xlsx")
sheet_obj = wb_obj.active 
row=sheet_obj.max_row
column=sheet_obj.max_column
#lists
Distance=[]
#setting in lists
for i in range(2, row+1):
    for h in range(2,column+1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        cell_obj2= sheet_obj.cell(row = 1, column = h)
        cell_obj3= sheet_obj.cell(row = i, column = h)
        appender=cell_obj.value+":"+str(cell_obj3.value)+":"+cell_obj2.value
        Distance.append(appender)
p.dump( Distance, open("Distances(the analyzed).protected","wb"))
